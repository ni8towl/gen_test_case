from tqdm import tqdm
import openpyxl

def copy_columns_between_excel_files(
    source_file: str,
    destination_file: str,
    source_sheet_name: str,
    destination_sheet_name: str,
    column_range: str,
    start_row: int,
    end_row: int,
    start_cell_in_destination: str
):
    """
    Copies specified columns from a source Excel file to a destination Excel file,
    starting from a specified cell in the destination sheet.

    Parameters:
        source_file (str): Path to the source Excel file.
        destination_file (str): Path to the destination Excel file.
        source_sheet_name (str): Name of the sheet in the source file.
        destination_sheet_name (str): Name of the sheet in the destination file.
        column_range (str): The columns to copy (e.g., "A:BL").
        start_row (int): The starting row to copy from in the source sheet.
        end_row (int): The ending row to copy from in the source sheet.
        start_cell_in_destination (str): The top-left cell where the range will be pasted (e.g., "C3").
    """
    # Load workbooks
    source_wb = openpyxl.load_workbook(source_file, data_only=True)
    destination_wb = openpyxl.load_workbook(destination_file)

    # Load sheets
    source_sheet = source_wb[source_sheet_name]
    destination_sheet = destination_wb[destination_sheet_name]

    # Parse the start cell in the destination
    start_row_dest, start_col_dest = openpyxl.utils.cell.coordinate_to_tuple(start_cell_in_destination)

    # Get the starting and ending columns from the column range
    start_col, end_col = column_range.split(":")
    start_col_index = openpyxl.utils.column_index_from_string(start_col)
    end_col_index = openpyxl.utils.column_index_from_string(end_col)

    # Calculate total cells to process for progress bar
    total_cells = (end_col_index - start_col_index + 1) * (end_row - start_row + 1)

    # Copy columns with a progress bar
    with tqdm(total=total_cells, desc="Copying columns", unit="cell") as pbar:
        for col_index in range(start_col_index, end_col_index + 1):
            dest_col_index = start_col_dest + (col_index - start_col_index)
            col_letter = openpyxl.utils.get_column_letter(col_index)
            dest_col_letter = openpyxl.utils.get_column_letter(dest_col_index)

            # Copy each cell in the specified row range
            for row in range(start_row, end_row + 1):
                source_value = source_sheet[f"{col_letter}{row}"].value
                dest_row = start_row_dest + (row - start_row)
                destination_sheet[f"{dest_col_letter}{dest_row}"].value = source_value

                # Update progress bar
                pbar.update(1)

    # Save the destination file
    destination_wb.save(destination_file)
    print("Columns copied successfully starting from the specified cell!")