import pandas as pd
from itertools import product
import tkinter as tk
from tkinter import simpledialog, filedialog
from prettytable import from_csv
from openpyxl import load_workbook
from excel_utils_v2 import copy_columns_between_excel_files
from tkinter import IntVar, Radiobutton
import json
from PIL import Image, ImageTk
import os

def get_user_input_with_image():
    # Create a hidden root window
    root = tk.Tk()
    root.overrideredirect(True)  # Hide the root window completely
    root.geometry("0x0+0+0")     # Make the root window invisible

    # Create a Toplevel dialog
    input_dialog = tk.Toplevel(root)  # Attach to the hidden root
    input_dialog.title("Enter Number of Inputs")

    # Add a label for instructions
    label_instruction = tk.Label(
        input_dialog,
        text="Select the number of inputs (1 to 10):",
        font=("Helvetica", 14, "bold")
    )

    label_instruction.pack(pady=10)

    # Add a sliding bar (scale) for user input
    input_var = tk.IntVar(value=1)  # Default value is 1
    scale_input = tk.Scale(
        input_dialog,
        variable=input_var,
        from_=1,
        to=10,
        orient="horizontal",
        length=400,
        tickinterval=1
    )
    scale_input.pack(pady=10)

    # Add an image below the sliding bar
    script_dir = os.path.dirname(os.path.abspath(__file__))
    image_path = os.path.join(script_dir, "Images", "Inputs.PNG")
    try:
        img = Image.open(image_path)
        img = img.resize((550, 700))
        img_tk = ImageTk.PhotoImage(img)

        label_image = tk.Label(input_dialog, image=img_tk)
        label_image.image = img_tk  # Keep a reference to avoid garbage collection
        label_image.pack(pady=10)
    except Exception as e:
        label_error = tk.Label(input_dialog, text=f"Error loading image: {e}")
        label_error.pack(pady=10)

    # Add a submit button
    def on_submit():
        input_dialog.result = str(input_var.get())  # Convert the selected value to a string
        input_dialog.destroy()

    button_submit = tk.Button(input_dialog, text="Submit", command=on_submit)
    button_submit.pack(pady=10)

    # Center the dialog and wait for it to close
    input_dialog.geometry("600x900")  # Adjust dimensions as needed
    input_dialog.transient(root)  # Attach to the hidden root
    input_dialog.grab_set()       # Block interactions with other windows
    root.wait_window(input_dialog)  # Wait until the dialog is closed

    # Clean up the hidden root window
    root.destroy()

    return input_dialog.result

# Define the function with default logic (can be overridden by user input)
def Output1(*inputs):
    # Map Input1, Input2, ..., InputN to the respective values
    local_scope = {f"Input{i+1}": inputs[i] for i in range(len(inputs))}
    return eval(user_logic, {}, local_scope)

def choose_test_type():
    """
    Opens a pop-up GUI with radio buttons for the user to select between two test types.
    Blocks further execution until the user confirms their selection.

    Returns:
        int: The value corresponding to the selected test type.
            1 -> SPC (True/False)
            2 -> DPC (OPEN/CLOSED)
    """
    # Create a hidden root window
    root = tk.Tk()
    root.overrideredirect(True)  # Hide the root window completely
    root.geometry("0x0+0+0")     # Make the root window invisible

    # Create a Toplevel dialog
    type_dialog = tk.Toplevel(root)  # Attach to the hidden root
    type_dialog.title("Choose Test Type")

    # Add a label for instructions
    label_instruction = tk.Label(
        type_dialog,
        text="Select the test type:\nChoose one of these options:",
        font=("Arial", 12)
    )
    label_instruction.pack(pady=10)

    # Add radio buttons for selecting test type
    test_type_var = tk.IntVar(value=1)  # Default value is 1
    radio_spc = tk.Radiobutton(
        type_dialog,
        text="1 -> SPC (True/False)",
        variable=test_type_var,
        value=1,
        font=("Arial", 10)
    )
    radio_spc.pack(pady=5)
    radio_dpc = tk.Radiobutton(
        type_dialog,
        text="2 -> DPC (OPEN/CLOSED). For testing switchgear.\n(Intermediate and Bad are not included)",
        variable=test_type_var,
        value=2,
        font=("Arial", 10)
    )
    radio_dpc.pack(pady=5)

    # Add a submit button
    def on_submit():
        type_dialog.result = test_type_var.get()
        type_dialog.destroy()

    button_submit = tk.Button(type_dialog, text="Submit", command=on_submit)
    button_submit.pack(pady=20)

    # Center the dialog and wait for it to close
    type_dialog.geometry("400x300")  # Adjust dimensions as needed
    type_dialog.transient(root)  # Attach to the hidden root
    type_dialog.grab_set()       # Block interactions with other windows
    root.wait_window(type_dialog)  # Wait until the dialog is closed

    # Clean up the hidden root window
    root.destroy()

    # Validation
    test_type = type_dialog.result
    if test_type in [1, 2]:
        print(f"User selected: {test_type} ({'SPC (True/False)' if test_type == 1 else 'DPC (OPEN/CLOSED)'})")
        return test_type
    else:
        print("No valid selection made. Returning None.")
        return None

def get_user_logic(num_inputs):
    # Create a hidden root window
    root = tk.Tk()
    root.overrideredirect(True)  # Hide the root window completely
    root.geometry("0x0+0+0")     # Make the root window invisible

    # Create a custom dialog window
    logic_dialog = tk.Toplevel(root)  # Attach to the hidden root
    logic_dialog.title("Input Logic Expression")

    # Default logic string
    default_logic = " or ".join([f"Input{i+1}" for i in range(num_inputs)])

    # Add instructions label
    label_instruction = tk.Label(
        logic_dialog,
        text=f"Enter a logic expression using these variables:\n\n"
             f"You can use logic operators 'and', 'or', 'not'\n\n"
             f"{', '.join([f'Input{i+1}' for i in range(num_inputs)])}\n\n"
             f"Default Example: {default_logic}\n",
        justify="left",
        font=("Helvetica", 14, "bold")
    )
    label_instruction.pack(pady=10)

    # Add an entry widget for user input
    input_var = tk.StringVar(value=default_logic)
    entry_input = tk.Entry(logic_dialog, textvariable=input_var, width=100)
    entry_input.pack(pady=10)

    # Add an image below the input field
    script_dir = os.path.dirname(os.path.abspath(__file__))
    image_path = os.path.join(script_dir, "Images", "Logic.PNG")
    try:
        img = Image.open(image_path)
        img = img.resize((600, 250))
        img_tk = ImageTk.PhotoImage(img)

        label_image = tk.Label(logic_dialog, image=img_tk)
        label_image.image = img_tk  # Keep a reference to avoid garbage collection
        label_image.pack(pady=10)
    except Exception as e:
        label_error = tk.Label(logic_dialog, text=f"Error loading image: {e}")
        label_error.pack(pady=10)

    # Add a submit button
    def on_submit():
        logic_dialog.result = input_var.get()
        logic_dialog.destroy()

    button_submit = tk.Button(logic_dialog, text="Submit", command=on_submit)
    button_submit.pack(pady=10)

    # Center the dialog and wait for it to close
    logic_dialog.geometry("1100x580")  # Adjust dimensions as needed
    logic_dialog.transient(root)  # Attach to the hidden root
    logic_dialog.grab_set()       # Block interactions with other windows
    root.wait_window(logic_dialog)  # Wait until the dialog is closed

    # Clean up the hidden root window
    root.destroy()

    return logic_dialog.result or default_logic

# Function to generate a truth table
def truth_table(f, num_inputs):
    variable_names = [f"Input{i+1}" for i in range(num_inputs)]
    values = [list(x) + [f(*x)] for x in product([False, True], repeat=num_inputs)]
    return pd.DataFrame(values, columns=variable_names + [f.__name__])

# Select destination Excel file
def select_xlsx_file():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="Select an Excel file",
        filetypes=[("Excel files", "*.xlsx")],
        defaultextension=".xlsx"
    )
    root.destroy()
    return file_path

# Main logic
if __name__ == "__main__":
    # Initialize Tkinter
    # root = tk.Tk()
    # root.withdraw()

    # Get user input with the custom dialog
    user_input = get_user_input_with_image()
    if not user_input or not user_input.isdigit() or int(user_input) <= 0:
        print("Invalid input. Please enter a valid number.")
        # root.destroy()
        exit()

    num_inputs = int(user_input)  # Convert the validated input to an integer

    print(f"User entered: {num_inputs}")

    # Call the choose_test_type function
    test_type = choose_test_type()

    # Handle the returned value
    if test_type == 1:
        print("Proceeding with SPC (True/False) tests...")
    elif test_type == 2:
        print("Proceeding with DPC (OPEN/CLOSED) tests...")
    else:
        print("No valid test type selected.")
        root.destroy()
        exit()

    # Save the selected test_type to a JSON file
    with open("test_type.json", "w") as f:
        json.dump({"test_type": test_type}, f)
    print(f"test_type saved to test_type.json: {test_type}")

    # Get user-defined logic
    user_logic = get_user_logic(num_inputs)

    # Generate and save truth table
    truth_table_df = truth_table(Output1, num_inputs)
    truth_table_df.to_csv('out.csv', index=False)

    # Update truth table based on test_type
    if test_type == 1:
        # Handle SPC (True/False) replacement
        for i in range(1, num_inputs + 1):
            col = f"Input{i}"
            truth_table_df[col] = truth_table_df[col].astype(str)
            truth_table_df[col] = truth_table_df[col].replace(["False", "FALSE"], "false")
            truth_table_df[col] = truth_table_df[col].replace(["True", "TRUE"], "true")
    elif test_type == 2:
        # Handle DPC (OPEN/CLOSED) replacement
        for i in range(1, num_inputs + 1):
            col = f"Input{i}"
            truth_table_df[col] = truth_table_df[col].astype(str)
            truth_table_df[col] = truth_table_df[col].replace(["False", "FALSE"], "CLOSED")
            truth_table_df[col] = truth_table_df[col].replace(["True", "TRUE"], "OPEN")

    # Save the updated truth table
    truth_table_df.to_csv("out_updated.csv", index=False)
    print("Updated CSV saved as out_updated.csv")

    # Print the truth table
    with open("out_updated.csv") as fp:
        mytable = from_csv(fp)
    print(mytable)

    # Transpose and save to Excel
    df_transposed = truth_table_df.transpose()
    df_transposed.to_excel("output_file_transposed.xlsx", index=False, header=False)
    print("Data successfully saved to output_file_transposed.xlsx")
    print(df_transposed)

    # Copy columns to destination Excel file
    destination_file = select_xlsx_file()
    if destination_file:
        source_file = "output_file_transposed.xlsx"
        column_range = "A:FAN"

        # Clear all cell values from row 3 downwards in the destination file
        print("Clearing data from row 3 downwards in the destination file...")
        workbook = load_workbook(destination_file)
        sheet = workbook["Test Steps"]
        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.value = None
        workbook.save(destination_file)
        print("Data cleared from row 3 downwards.")

        copy_columns_between_excel_files(
            source_file=source_file,
            destination_file=destination_file,
            source_sheet_name="Sheet1",
            destination_sheet_name="Test Steps",
            column_range=column_range,
            start_row=1,
            end_row=len(truth_table_df) + 1,
            start_cell_in_destination="C3"
        )
        print("Completed copying the specified columns starting from C3!")

        # Open the destination workbook to modify
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(destination_file, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            sheet = workbook["Test Steps"]

            # Find the last row with actual data in column C
            last_row = max(
                row[0].row for row in sheet.iter_rows(min_col=3, max_col=3, min_row=1, max_row=sheet.max_row)
                if row[0].value is not None
            )
            print(f"Inserting 'ASSESS' into Column B at row {last_row}")

            # Insert 'ASSESS' into the appropriate cell
            sheet[f"B{last_row}"] = "ASSESS"

            # Save changes
            workbook.save(destination_file)

    # Clean up the root window
    # root.destroy()
